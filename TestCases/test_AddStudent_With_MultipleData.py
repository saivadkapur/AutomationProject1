import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_students():
    #API
    add_url = "http://www.thetestingworldapi.com/api/studentsDetails"
    file = open('D:/AutomationSampleDataFiles/RequestJson.json', 'r')
    json_request = json.loads(file.read())

    #EXCEL
    #you need to import 'OPENPYXL' library to access excel files to deal with the multiple data
    #'openpyxl.laod_workbook() is used to load the excel files
    wk = openpyxl.load_workbook('D:/AutomationSampleDataFiles/StudentsMultipleData.xlsx')
    #now enter the sheet number of the excel file on which you want to work on and holding multiple data in it
    sh = wk['Sheet1']
    #below line of code will fetch the total no of rows exist in the excel file
    rows = sh.max_row

    for i in range(2,rows+1):
        #Below are the lines of code used to create an object for each column in the excel
        cell_first_name = sh.cell(row=i, column=1)
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)
        #Now read the values from each row and for each element and pass through values
        #into the json body in here the json request is stored in RequestJson.json file
        #json_request['KeyName_from the request body'] = cell_first_name.value get the value for the key
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_mid_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value
        response = requests.post(add_url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201







