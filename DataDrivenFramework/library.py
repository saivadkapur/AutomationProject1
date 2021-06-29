#This module hold the common code
#like reading a excel file, fetching max no. of rows and columns in the excel file

import requests
import json
import jsonpath
import openpyxl

class Common:
    #Create a constructor to load the excel file and read the excel sheet
    def __init__(self, ExcelFilePath, SheetNumber):
        global wk
        global sh
        wk = openpyxl.load_workbook(ExcelFilePath)
        sh = wk[SheetNumber]

    #Method to fetch no of rows in the excel file
    def fetch_noof_rows(self):
        rows = sh.max_row
        return rows

    #Method to fetch no of columns in the excel sheet
    def fetch_noof_columns(self):
        columns = sh.max_column
        return columns

    #Method to fetch the value from the header of the excel file i.e each column value from the first row
    #of the excel and those will be the key values from the request body
    def fetch_keyname_jsonreqbody(self):
        c = sh.max_column
        #create a list to store the column names('first_name', 'middle_name','last_name', 'date_of_birth')
        li = []
        for i in range(1, c+1):
            cell = sh.cell(row=1, column=i)
            li.insert(i-1,cell.value)

        return li

    #Method to update json request body with the multiple data
    def update_requestbodywith_data(self,rowNumber,jsonRequest,keyList):
        c = sh.max_column
        for i in range (1, c+1):
            cell = sh.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value #jsonRequest[keyList[0]] =

        return jsonRequest


